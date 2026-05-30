import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from io import BytesIO
import re

▽▽▽ ここを新しく差し替えます（最後を html に修正しました） ▽▽▽
st.markdown(
    """
    <style>
    /* ① アップロードエリアの点線枠を真っ黒にして目立たせる */
    [data-testid="stFileUploaderDropzone"] {
        border: 2px dashed #000000 !important;
        background-color: #f8f9fa !important;
    }
    /* ② 点線枠の中の文字や案内（Drag and drop...など）を真っ黒＆太字にする */
    [data-testid="stFileUploaderDropzone"] {
        color: #000000 !important;
        font-weight: bold !important;
    }
    /* ③ 画面上のすべての通常の文字をハッキリとした黒にする */
    .stMarkdown, p, label {
        color: #000000 !important;
    }
    /* ④ タイトルをさらに大きく太く目立たせる */
    div[data-testid="stFileUploader"] label p {
        font-size: 1.3rem !important;
        font-weight: bold !important;
        color: #000000 !important;
    }
    </style>
    """,
    unsafe_allow_html=True  # 💡ここを「html」に直しました！
)
# △△△ ここまで △△△

st.title("発注明細リスト 自動加工ツール")

uploaded_file = st.file_uploader("Excelファイルをアップロード", type=["xlsx"])

if uploaded_file:
    st.info("処理を実行中… 少しお待ちください")

    # =========================
    # ① 元データ読み込み
    # =========================
    df = pd.read_excel(uploaded_file)

    # =========================
    # ② 不要な支店を削除
    # =========================
    df = df[~df["売上支店"].isin(["(支店)リフォーム", "(支店)外構工事", "リフォーム部"])]

    # =========================
    # ③ 業者名の統一
    # =========================
    replace_rules = {
        r"㈱丸増ﾍﾞﾆﾔ商会.*": "㈱丸増ﾍﾞﾆﾔ商会",
        r"中部ﾎｰﾑｻｰﾋﾞｽ㈱.*": "中部ﾎｰﾑｻｰﾋﾞｽ㈱",
        r"㈱ｺﾆｼ.*": "㈱ｺﾆｼ",
        r"㈱ｼﾞｭｰﾃｯｸ.*": "㈱ｼﾞｭｰﾃｯｸ",
        r"ｿﾆﾃｯｸ㈱.*": "ｿﾆﾃｯｸ㈱",
        r"ﾏﾃﾘｱﾙｴｰﾄﾞ㈱.*": "ﾏﾃﾘｱﾙｴｰﾄﾞ㈱",
        r"㈱ﾋﾟｺｲ.*": "㈱ﾋﾟｺｲ",
        r"ﾊﾟﾅｿﾆｯｸﾘﾋﾞﾝｸﾞ.*": "ﾊﾟﾅｿﾆｯｸﾘﾋﾞﾝｸﾞ",
        r"㈱丸八.*": "㈱丸八",
        r"岡田電気産業㈱.*": "岡田電気産業㈱",
        r"㈱山善.*": "㈱山善",
        r"ﾌｫｰﾑ断熱㈱.*": "ﾌｫｰﾑ断熱㈱",
        r"三協立山㈱.*": "三協立山㈱",
        r"㈱ﾔﾏｹﾝ.*": "㈱ﾔﾏｹﾝ",
        r"鳥居金属興業㈱.*": "鳥居金属興業㈱",
        r"杉田ｴｰｽ㈱.*": "杉田ｴｰｽ㈱"
    }

    for pattern, name in replace_rules.items():
        df["業者名"] = df["業者名"].str.replace(pattern, name, regex=True)

    # =========================
    # ④ 日付整形
    # =========================
    date_cols = [
        "契約日", "着工予定日", "着工実績日",
        "上棟予定日", "上棟実績日",
        "木完予定日", "木完実績日",
        "引渡予定日", "引渡実績日"
    ]

    for col in date_cols:
        df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%Y/%m/%d")
        df[col] = df[col].astype(str).str.replace(r"/0", "/", regex=True)

    # =========================
    # ⑤ 修正済みシート（メモリ上）
    # =========================
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="修正済み", index=False)

    # =========================
    # ⑥ 集計処理
    # =========================
    # 前半：項目ごとに集計する業者
    vendors_front = {
        "双日建材㈱": {
            "grouped": [
                "木材費(PC)", "ｷｯﾁﾝ(ﾀｶﾗ)", "ﾕﾆｯﾄﾊﾞｽ(TOTO)", "洗面化粧台(TOTO)",
                "建材神島(軒天材､幕板､付柱等)", "ｷｯﾁﾝ（LIXIL）", "洗面化粧台(LIXIL)",
                "ﾕﾆｯﾄﾊﾞｽ(LIXIL)", "外壁材(ﾆﾁﾊ)", "ｷｯﾁﾝ(永大産業)", "ｻｯｼ(LIXIL)", "ｴｱｺﾝ(材工)"
            ]
        },
        "阪和興業㈱":{
            "grouped": [
                "木材費(PC)", "ｻｯｼ(LIXIL)"
            ]
        },
        "SMB建材㈱": {
            "special": {"内装建材(NODA)": ["内装建具(NODA)", "内装建材(NODA)"]},
            "grouped": ["ｷｯﾁﾝ(ｸﾘﾅｯﾌﾟ)", "洗面化粧台(ｸﾘﾅｯﾌﾟ)", "ﾀｲﾙ工事", "外壁材(KMEW)", "照明(DAIKO)"]
        },
        "㈱丸増ﾍﾞﾆﾔ商会": {
            "special": {"内装建材(EIDAI)": ["内装建材(EIDAI)", "内装建具(EIDAI)"]},
            "grouped": [
                "断熱材(ｷｭｰﾜﾝﾎﾞｰﾄﾞ他)", "建材費(PB､合板､断熱)", "内装建材(朝日ウッドテック)",
                "ﾊｲﾍﾞｽﾄｳｯﾄﾞ", "内装建材(DAIKEN)", "気密部材(ｽﾀｲﾛ、ｳﾚﾀﾝ、ﾃｰﾌﾟ)"
            ]
        },
        "加藤ﾍﾞﾆﾔ㈱": {
            "grouped": [
                "ﾊｲﾍﾞｽﾄｳｯﾄﾞ", "建材費(PB､合板､断熱)", "気密部材(ｽﾀｲﾛ、ｳﾚﾀﾝ、ﾃｰﾌﾟ)", "内装建材(WOODONE)"
            ]
        },
        "㈱ｻﾝｺｰ(資材)": {
            "special": {"内装建材(EIDAI)": ["内装建材(EIDAI)", "内装建具(EIDAI)"]},
            "grouped": [
                "建材費(PB､合板､断熱)", "断熱材(ｷｭｰﾜﾝﾎﾞｰﾄﾞ他)", "ﾊｲﾍﾞｽﾄｳｯﾄﾞ",
                "内装建材(朝日ウッドテック)", "内装建材(DAIKEN)", "気密部材(ｽﾀｲﾛ、ｳﾚﾀﾝ、ﾃｰﾌﾟ)"
            ]
        },
        "㈱山大": {
            "special": {"内装建材(EIDAI)": ["内装建材(EIDAI)", "内装建具(EIDAI)"]},
            "grouped": [
                "外壁材(ﾆﾁﾊ)", "外壁材(KMEW)", "建材費(PB､合板､断熱)",
                "内装建材(朝日ウッドテック)", "ﾊｲﾍﾞｽﾄｳｯﾄﾞ",
                "気密部材(ｽﾀｲﾛ、ｳﾚﾀﾝ、ﾃｰﾌﾟ)", "断熱材(ｷｭｰﾜﾝﾎﾞｰﾄﾞ他)"
            ]
        },
        "㈱坂田建材": {
            "special": {
                "内装建材(EIDAI)": ["内装建材(EIDAI)", "内装建具(EIDAI)"],
                "内装建材(朝日ウッドテック)": ["内装建材(朝日ウッドテック)", "内装建具(朝日ウッドテック)"]
            },
            "grouped": [
                "断熱材(ｷｭｰﾜﾝﾎﾞｰﾄﾞ他)", "建材費(PB､合板､断熱)", "ﾊｲﾍﾞｽﾄｳｯﾄﾞ",
                "気密部材(ｽﾀｲﾛ、ｳﾚﾀﾝ、ﾃｰﾌﾟ)", "瓦・板金工事(工)", "雨樋(工)"
            ]
        },
        "村上木材㈱": {
            "special": {
                "内装建材(EIDAI)": ["内装建材(EIDAI)", "内装建具(EIDAI)"],
                "内装建材(朝日ウッドテック)": ["内装建材(朝日ウッドテック)", "内装建具(朝日ウッドテック)"],
                "内装建材(DAIKEN)": ["内装建材(DAIKEN)", "内装建具(DAIKEN)"],
                "内装建材(WOODONE)": ["内装建材(WOODONE)", "内装建具(WOODONE)"]
            },
            "grouped": [
                "断熱材(ｷｭｰﾜﾝﾎﾞｰﾄﾞ他)", "建材費(PB､合板､断熱)", "ﾊｲﾍﾞｽﾄｳｯﾄﾞ",
                "気密部材(ｽﾀｲﾛ、ｳﾚﾀﾝ、ﾃｰﾌﾟ)"
            ]
        },
        "中部ﾎｰﾑｻｰﾋﾞｽ㈱": {
            "special": {
                "内装建材(EIDAI)": ["内装建材(EIDAI)", "内装建具(EIDAI)"],
                "内装建材(朝日ウッドテック)": ["内装建材(朝日ウッドテック)", "内装建具(朝日ウッドテック)"],
                "外壁工事": ["外壁材(ﾆﾁﾊ)", "外壁張手間", "外壁材(KMEW)"]
            },
            "grouped": [
                "断熱材(ｷｭｰﾜﾝﾎﾞｰﾄﾞ他)", "建材費(PB､合板､断熱)", "ﾊｲﾍﾞｽﾄｳｯﾄﾞ"
            ]
        },
        "㈱ｺﾆｼ": {
            "special": {
                "内装建材(EIDAI)": ["内装建材(EIDAI)", "内装建具(EIDAI)"],
                "内装建材(朝日ウッドテック)": ["内装建材(朝日ウッドテック)", "内装建具(朝日ウッドテック)"],
                "内装建材(DAIKEN)": ["内装建材(DAIKEN)", "内装建具(DAIKEN)"],
                "内装建材(WOODONE)": ["内装建材(WOODONE)", "内装建具(WOODONE)"]
            },
            "grouped": [
                "断熱材(ｷｭｰﾜﾝﾎﾞｰﾄﾞ他)", "建材費(PB､合板､断熱)", "ﾊｲﾍﾞｽﾄｳｯﾄﾞ","気密部材(ｽﾀｲﾛ、ｳﾚﾀﾝ、ﾃｰﾌﾟ)"
            ]
        },
        "㈱ｼﾞｭｰﾃｯｸ": {"grouped": ["太陽光発電(材工)", "全館空調換気ｼｽﾃﾑ(ﾛｰﾔﾙ電機)", "給湯器"]},
        "ｿﾆﾃｯｸ㈱": {"grouped": ["木材費(金物)"]},
        "ﾏﾃﾘｱﾙｴｰﾄﾞ㈱": {"grouped": ["ｲﾝﾀｰﾎﾝ･照明･換気扇他", "分電盤", "屋外フード（材）", "火災報知器（材）", "ﾌｰﾄﾞ･火報等(材)"]},
        "㈱共ｼｮｳ": {"grouped": ["断熱材(ｸﾞﾗｽｳｰﾙ)", "建材費(点検口､換気材、水切材)", "物干し金物"]},
        "㈱ﾋﾟｺｲ": {"grouped": ["防虫･防蟻工事(材工)", "気密測定"]},
        "日精ﾌﾟﾗｽﾃｯｸ㈱": {"grouped": ["制震ｼｽﾃﾑ", "建材費(床断熱材)ﾌｪﾉﾊﾞﾎﾞｰﾄﾞ"]},
        "ﾊﾟﾅｿﾆｯｸﾘﾋﾞﾝｸﾞ": {"grouped": ["ｴｺｷｭｰﾄ(ﾊﾟﾅｿﾆｯｸ)"]},
        "㈱丸八": {"grouped": ["ﾕﾆｯﾄﾊﾞｽ(TOTO)", "洗面化粧台(TOTO)"]},
        "岡田電気産業㈱": {"grouped": ["照明(DAIKO)"]},
        "㈱山善": {"grouped": ["ｴｺｷｭｰﾄ(三菱)", "ｴｱｺﾝ(材工)"]}
    }

    # 後半：合計だけ出す業者
    vendors_back = [
        "日本制震ｼｽﾃﾑ㈱","Solvvy㈱","㈱篠原商店","㈱ｹｰ･ｴｲﾁ･ｹｰ",
        "㈱ｼｰ･ｴｽ･ﾗﾝﾊﾞｰ","ﾛｰﾔﾙ電機㈱","㈱富士通ｾﾞﾈﾗﾙ","㈱ｻﾑｼﾝｸﾞ",
        "ﾕｱｻｸｵﾋﾞｽ㈱","田村駒ｴﾝｼﾞﾆｱﾘﾝｸﾞ㈱","ﾌｫﾜｰﾄﾞ.H.S㈱","東京ｸﾞﾗｽﾛﾝ㈱",
        "ﾌｫｰﾑ断熱㈱","三協立山㈱","㈱ｵﾝﾀﾞ製作所","㈱ｱﾄﾞｳﾞｧﾝｸﾞﾙｰﾌﾟ",
        "㈱ﾔﾏｹﾝ","YKｱｸﾛｽ㈱","菊地合板木工㈱","㈲YSKｻﾎﾟｰﾄ",
        "鳥居金属興業㈱","㈱竹屋化学研究所","㈱関家具","協立ｴｱﾃｯｸ㈱",
        "杉田ｴｰｽ㈱","㈱ﾖﾈｷﾝ"
    ]

    def summarize_vendor(df, vendor, special_items=None, grouped_items=None):
        df_v = df[df["業者名"] == vendor]
        summary = {}

        if special_items:
            for name, items in special_items.items():
                summary[name] = df_v.loc[df_v["工種名"].isin(items), "金額"].sum()

        if grouped_items:
            for item in grouped_items:
                summary[item] = df_v.loc[df_v["工種名"] == item, "金額"].sum()

        exclude = []
        if special_items:
            for items in special_items.values():
                exclude += items
        if grouped_items:
            exclude += grouped_items

        summary["その他"] = df_v.loc[~df_v["工種名"].isin(exclude), "金額"].sum()
        summary[f"{vendor} 合計"] = df_v["金額"].sum()

        return summary

    all_summaries = []

    # 前半の業者（項目別集計）
    for vendor, rules in vendors_front.items():
        summary = summarize_vendor(
            df,
            vendor,
            special_items=rules.get("special"),
            grouped_items=rules.get("grouped")
        )
        all_summaries.extend(summary.items())

    # 後半の業者（合計のみ）
    for vendor in vendors_back:
        total = df.loc[df["業者名"] == vendor, "金額"].sum()
        all_summaries.append((f"{vendor} 合計", total))

    summary_df = pd.DataFrame(all_summaries, columns=["工種名", "合計金額"])

    # =========================
    # ⑧ 集計シートを Excel に追加
    # =========================
    with pd.ExcelWriter(output, engine="openpyxl", mode="a") as writer:
        summary_df.to_excel(writer, sheet_name="集計", index=False)

    # =========================
    # ⑨ 書式設定（合計行を太字＋枠線＋水色）
    # =========================
    output.seek(0)
    wb = load_workbook(output)
    ws = wb["集計"]

    bold_font = Font(bold=True)
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    fill_blue = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        cell_name, cell_value = row
        if cell_name.value and str(cell_name.value).endswith("合計"):
            row[0].font = bold_font
            row[1].font = bold_font
            row[0].border = border_style
            row[1].border = border_style
            row[0].fill = fill_blue
            row[1].fill = fill_blue

    # 書式反映後の Excel を再生成
    output2 = BytesIO()
    wb.save(output2)
    output2.seek(0)

    # =========================
    # ⑩ ダウンロードボタン
    # =========================
    st.success("自動加工が完了しました！ボタンを押してダウンロードしてください。")
    st.download_button(
        label="加工済みExcel（書式付き）をダウンロード",
        data=output2.getvalue(),
        file_name="発注明細_加工済み.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )